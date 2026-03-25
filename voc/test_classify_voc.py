"""RED-phase tests for VOC classification script."""

# pyright: reportMissingImports=false, reportMissingModuleSource=false, reportImplicitRelativeImport=false, reportAttributeAccessIssue=false, reportOptionalMemberAccess=false
import json
import copy
import importlib
import xml.etree.ElementTree as ET
from pathlib import Path
import re
import shutil
import subprocess
import sys
import tempfile
import unittest
from unittest.mock import patch, MagicMock
from io import BytesIO
from zipfile import ZipFile

import openpyxl
import pytest
from openpyxl.styles import Border, Side


VALID_CATEGORIES = {
    "반응_문의",
    "반응_교환",
    "반응_구해요",
    "반응_정보",
    "반응_거래",
    "반응_마감문의",
    "반응_단순반응",
    "반응_제품칭찬",
    "반응_제품비판",
    "반응_일반감상",
    "반응_구매추천",
    "반응_구매비추천",
}


def _build_base_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026 list"
    ws.cell(4, 7).value = "주요 내용"
    ws.cell(4, 10).value = "제목"
    return wb, ws


def _add_card(ws, start_col, link, category="반응_제품반응"):
    ws.cell(2, start_col).value = "Voice Of Customer"
    ws.cell(4, start_col).value = "사이트"
    ws.cell(4, start_col + 3).value = "링크"
    ws.cell(4, start_col + 4).value = link
    ws.cell(5, start_col).value = "주요 이슈"
    ws.cell(5, start_col + 1).value = category


def _mock_http_response(payload):
    response = MagicMock()
    response.__enter__.return_value.read.return_value = json.dumps(payload).encode(
        "utf-8"
    )
    return response


def _http_error(url):
    http_error_cls = __import__("urllib.error", fromlist=["HTTPError"]).HTTPError
    return http_error_cls(url, 500, "mock error", hdrs=None, fp=None)


def _pipeline_callable(module):
    return getattr(module, "run_classification", None) or getattr(
        module, "classify_workbook", None
    )


def _real_input_workbook() -> Path:
    candidates = [
        Path("input/2026더캐리VOC(리포트파일)_260316~260322.xlsx"),
        Path("input/2026더캐리VOC(리포트파일)_260309~260315.xlsx"),
    ]
    for path in candidates:
        if path.exists():
            return path
    pytest.skip("real input workbook not available")
    raise AssertionError("unreachable")


def _xlsx_part_counts(path: Path) -> tuple[int, int]:
    with ZipFile(path) as zf:
        media = [name for name in zf.namelist() if name.startswith("xl/media/")]
        drawings = [name for name in zf.namelist() if name.startswith("xl/drawings/")]
    return len(media), len(drawings)


def _xlsx_member_crc_map(path: Path) -> dict[str, int]:
    with ZipFile(path) as zf:
        return {item.filename: item.CRC for item in zf.infolist()}


def _changed_zip_parts(source: Path, output: Path) -> set[str]:
    src = _xlsx_member_crc_map(source)
    out = _xlsx_member_crc_map(output)
    names = set(src) | set(out)
    return {name for name in names if src.get(name) != out.get(name)}


def _sheet_max_col_map(path: Path) -> dict[str, int]:
    wb = openpyxl.load_workbook(path)
    return {ws.title: ws.max_column for ws in wb.worksheets}


def _build_legacy_baseline_updates(src: Path):
    import classify_voc

    wb = openpyxl.load_workbook(src)
    targets = classify_voc.read_target_rows(wb)
    if not targets:
        pytest.skip("fixture has no target rows")
    first_target = targets[0]
    return classify_voc._build_sheet_updates(
        first_target["row"],
        first_target.get("link"),
        "반응_일반감상",
        classify_voc.build_card_index(wb),
    )


def _save_value_only_patch(source_path: Path, output_path: Path, updates):
    import classify_voc

    by_sheet: dict[str, dict[str, str]] = {}
    for update in updates:
        by_sheet.setdefault(update["sheet_name"], {})[update["cell_ref"]] = update[
            "value"
        ]

    with ZipFile(source_path, "r") as source_zip:
        sheet_paths = classify_voc._sheet_xml_paths(source_zip)
        patched_entries: dict[str, bytes] = {}
        for sheet_name, sheet_updates in by_sheet.items():
            sheet_path = sheet_paths.get(sheet_name)
            if not sheet_path:
                continue
            patched_entries[sheet_path] = classify_voc._patch_sheet_xml(
                source_zip.read(sheet_path), sheet_updates
            )

        with ZipFile(output_path, "w") as out_zip:
            out_zip.comment = source_zip.comment
            for item in source_zip.infolist():
                data = patched_entries.get(item.filename)
                if data is None:
                    data = source_zip.read(item.filename)
                out_zip.writestr(item, data)


def _apply_card_outline(ws, start_col: int):
    medium = Side(style="medium", color="000000")
    for row_idx in range(2, 36):
        for col_idx in range(start_col, start_col + 6):
            cell = ws.cell(row_idx, col_idx)
            if row_idx == 2:
                cell.border = Border(
                    top=medium,
                    left=cell.border.left,
                    right=cell.border.right,
                    bottom=cell.border.bottom,
                )
            if row_idx == 35:
                cell.border = Border(
                    bottom=medium,
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                )
            if col_idx == start_col:
                cell.border = Border(
                    left=medium,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )
            if col_idx == start_col + 5:
                cell.border = Border(
                    right=medium,
                    left=cell.border.left,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )


def _supported_workbook_contract():
    return importlib.import_module("xlsx_template_contract").load_template_contract()


def _contract_workbook_hashes(contract):
    family = contract["supported_workbook_family"]
    hashes = set()
    if family.get("workbook_xml_sha256"):
        hashes.add(family["workbook_xml_sha256"])
    hashes.update(family.get("workbook_xml_sha256_allowlist", []))
    return hashes


def _contract_sheet_state_variants(contract):
    topology = contract["preserved_topology"]
    variants = topology.get("sheet_states_variants")
    if variants:
        return variants
    return [topology["sheet_states"]]


def _contract_linked_card_target_variants(contract):
    writable_surface = contract["writable_surface"]
    variants = [writable_surface["linked_card_targets"]]
    variants.extend(writable_surface.get("linked_card_targets_variants", []))
    return variants


def _xlsx_package_diff_validator_cmd(template: Path, output: Path) -> list[str]:
    repo_root = Path(__file__).resolve().parent
    return [
        sys.executable,
        str(repo_root / "scripts/check_xlsx_package_diff.py"),
        "--template",
        str(template),
        "--output",
        str(output),
        "--allowlist",
        str(repo_root / ".sisyphus/contracts/xlsx-allowlist.json"),
    ]


def _rewrite_zip_member(path: Path, member_name: str, mutate):
    with tempfile.TemporaryDirectory() as tmpdir:
        patched = Path(tmpdir) / "patched.xlsx"
        with ZipFile(path, "r") as src_zip, ZipFile(patched, "w") as out_zip:
            out_zip.comment = src_zip.comment
            for item in src_zip.infolist():
                data = src_zip.read(item.filename)
                if item.filename == member_name:
                    data = mutate(data)
                out_zip.writestr(item, data)
        patched.replace(path)


def _mutate_first_non_writable_cell(
    xml_bytes: bytes, writable_cells: set[str]
) -> bytes:
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    root = ET.fromstring(xml_bytes)
    for cell in root.findall(f".//{{{ns}}}c"):
        cell_ref = cell.attrib.get("r")
        if not cell_ref or cell_ref in writable_cells:
            continue
        if cell.find(f"{{{ns}}}f") is not None:
            continue
        value_node = cell.find(f"{{{ns}}}v")
        if value_node is not None:
            value_node.text = (value_node.text or "") + "__drift__"
            return ET.tostring(root, encoding="utf-8", xml_declaration=True)
        inline_text = cell.find(f"{{{ns}}}is/{{{ns}}}t")
        if inline_text is not None:
            inline_text.text = (inline_text.text or "") + "__drift__"
            return ET.tostring(root, encoding="utf-8", xml_declaration=True)
    raise AssertionError("failed to locate a non-writable cell for mutation")


def _updates_by_sheet(updates):
    grouped: dict[str, dict[str, str]] = {}
    for update in updates:
        grouped.setdefault(update["sheet_name"], {})[update["cell_ref"]] = update[
            "value"
        ]
    return grouped


class TestVocClassifier(unittest.TestCase):
    def test_rule_classify_exchange(self):
        import classify_voc

        assert classify_voc.rule_classify("[교환] 사이즈 바꿔요") == "반응_교환"

    def test_rule_classify_seeking(self):
        import classify_voc

        assert classify_voc.rule_classify("[구해요] 130 구해요") == "반응_구해요"

    def test_rule_classify_no_match(self):
        import classify_voc

        assert classify_voc.rule_classify("오늘 배송 빠르네요") is None

    def test_rule_classify_info_keywords(self):
        import classify_voc

        assert (
            classify_voc.rule_classify("푸마 에이치스트릿 품절풀렸어요") == "반응_정보"
        )
        assert classify_voc.rule_classify("캐리마켓 최대 80% 할인 핫딜") == "반응_정보"

    def test_rule_classify_info_price_pattern_only(self):
        import classify_voc

        assert classify_voc.rule_classify("캐리마켓 14,310원 핫딜") == "반응_정보"
        assert (
            classify_voc.rule_classify(
                "무신사) 푸마 키즈 스피드캣 발렛 키즈 79000원 무배"
            )
            == "반응_문의"
        )
        assert (
            classify_voc.rule_classify("원픽이긴 한데 구매처 아시는분") == "반응_문의"
        )

    def test_rule_priority_keeps_seeking_before_info(self):
        import classify_voc

        assert (
            classify_voc.rule_classify("[구해요] 150 구해요 79,000원") == "반응_구해요"
        )

    def test_rule_classify_trade_completion(self):
        import classify_voc

        assert (
            classify_voc.rule_classify("(완료)베베드피노 몬치치 오버롤 90사이즈")
            == "반응_거래"
        )

    def test_rule_classify_finish_question(self):
        import classify_voc

        assert (
            classify_voc.rule_classify("베베드피노 원래 마감이 좀 깔끔치 못한가요?")
            == "반응_마감문의"
        )

    def test_rule_classify_simple_reaction(self):
        import classify_voc

        assert (
            classify_voc.rule_classify("가방 왜 가벼워야 하는지 이제 알겠네요..ㅠ")
            == "반응_단순반응"
        )

    def test_rule_classify_purchase_success_as_info(self):
        import classify_voc

        assert (
            classify_voc.rule_classify("푸마키즈 에이치스트릿 구매 성공👍")
            == "반응_정보"
        )

    def test_rule_classify_does_not_turn_purchase_question_into_info(self):
        import classify_voc

        assert (
            classify_voc.rule_classify("푸마 스피드캣 고 구매처 아시는분 ㅜ")
            == "반응_문의"
        )

    @patch("classify_voc.urllib.request.urlopen")
    def test_fetch_text_uses_response_charset(self, mock_urlopen):
        import classify_voc

        response = MagicMock()
        response.__enter__.return_value.read.return_value = "재입고 되었어요".encode(
            "cp949"
        )
        response.__enter__.return_value.headers.get_content_charset.return_value = (
            "ms949"
        )
        mock_urlopen.return_value = response

        text = classify_voc._fetch_text("https://example.com/post/1")

        assert text == "재입고 되었어요"

    def test_extract_article_text_rejects_low_quality_chrome_text(self):
        import classify_voc

        html = """
        <div class="article viewer">
          <p>게시판 목록 바로가기</p>
          <p>본문 바로가기</p>
          <p>내소식이 없습니다.</p>
        </div>
        """

        assert classify_voc.extract_article_text(html) is None

    def test_build_classification_prompt(self):
        import classify_voc

        prompt = classify_voc.build_classification_prompt("이 제품 어디서 구매하나요")
        for category in VALID_CATEGORIES:
            assert category in prompt
        assert any(
            token in prompt for token in ["궁금", "교환", "구해", "가격", "성분"]
        )
        assert any(token in prompt for token in ["무배", "리셀", "핫딜", "신상"])
        assert "반응_기타" not in prompt

    def test_build_body_classification_prompt(self):
        import classify_voc

        prompt = classify_voc.build_body_classification_prompt(
            "추천할만한지, 마감은 어떤지 궁금해요"
        )
        for category in VALID_CATEGORIES:
            assert category in prompt
        assert "반응_기타" not in prompt
        assert "반응_구매추천" in prompt
        assert "반응_구매비추천" in prompt

    def test_github_gpt41_adapter_request_format(self):
        import classify_voc

        adapter = classify_voc.GithubModelsAdapter(
            api_key="test-key", model="openai/gpt-4.1"
        )
        body = adapter.build_request_body("테스트 제목")

        assert body["model"] == "openai/gpt-4.1"
        assert "messages" in body
        assert isinstance(body["messages"], list)
        assert len(body["messages"]) >= 1

    def test_github_gpt41_adapter_parse_response(self):
        import classify_voc

        adapter = classify_voc.GithubModelsAdapter(
            api_key="test-key", model="openai/gpt-4.1"
        )
        mock_response = {"choices": [{"message": {"content": "반응_문의"}}]}

        result = adapter.parse_response(mock_response)
        assert result == "반응_문의"

    def test_github_gpt4o_adapter_request_format(self):
        import classify_voc

        adapter = classify_voc.GithubModelsAdapter(
            api_key="test-key", model="openai/gpt-4o"
        )
        body = adapter.build_request_body("테스트 제목")

        assert body["model"] == "openai/gpt-4o"
        assert "messages" in body

    def test_github_gpt4o_adapter_parse_response(self):
        import classify_voc

        adapter = classify_voc.GithubModelsAdapter(
            api_key="test-key", model="openai/gpt-4o"
        )
        mock_response = {"choices": [{"message": {"content": "반응_교환"}}]}

        result = adapter.parse_response(mock_response)
        assert result == "반응_교환"

    def test_validate_classification_result(self):
        import classify_voc

        assert classify_voc.validate_result("반응_문의") == "반응_문의"
        assert classify_voc.validate_result("반응_교환") == "반응_교환"
        assert classify_voc.validate_result("반응_구해요") == "반응_구해요"
        assert classify_voc.validate_result("반응_정보") == "반응_정보"
        assert classify_voc.validate_result("반응_거래") == "반응_거래"
        assert classify_voc.validate_result("반응_마감문의") == "반응_마감문의"
        assert classify_voc.validate_result("반응_단순반응") == "반응_단순반응"
        assert classify_voc.validate_result("반응_제품칭찬") == "반응_제품칭찬"
        assert classify_voc.validate_result("반응_제품비판") == "반응_제품비판"
        assert classify_voc.validate_result("반응_일반감상") == "반응_일반감상"
        assert classify_voc.validate_result("반응_구매추천") == "반응_구매추천"
        assert classify_voc.validate_result("반응_구매비추천") == "반응_구매비추천"
        assert classify_voc.validate_result("반응_기타") == "반응_단순반응"
        assert classify_voc.validate_result("알 수 없음") == "반응_단순반응"
        assert classify_voc.validate_result("") == "반응_단순반응"
        assert classify_voc.validate_result("긍정_제품후기") == "반응_정보"
        assert classify_voc.validate_result("비추천") == "반응_구매비추천"
        assert classify_voc.validate_result("(완료) 거래") == "반응_거래"
        assert classify_voc.validate_result("마감이 별로인가요") == "반응_마감문의"

    def test_preprocess_title(self):
        import classify_voc

        newline_case = classify_voc.preprocess_title("제목\n부제목")
        assert newline_case in ("제목 부제목", "제목")

        assert classify_voc.preprocess_title("  공백 포함  ") == "공백 포함"

        try:
            empty_result = classify_voc.preprocess_title("")
            assert empty_result is None
        except ValueError:
            pass

        try:
            none_result = classify_voc.preprocess_title(None)
            assert none_result is None
        except TypeError:
            pass

    def test_build_card_index_includes_duplicate_links(self):
        import classify_voc

        wb, _ = _build_base_workbook()
        ws1 = wb.create_sheet("IB")
        ws2 = wb.create_sheet("KR")
        link = "https://example.com/post/1"
        _add_card(ws1, 2, link)
        _add_card(ws2, 14, link)

        card_index = classify_voc.build_card_index(wb)

        assert link in card_index
        assert len(card_index[link]) == 2

    def test_extract_article_text_removes_notice(self):
        import classify_voc

        html = """
        <div class="article viewer">
          <p>회원간의 거래 분쟁에 대한 공론화 금지</p>
          <p>가격이 얼마인가요?</p>
        </div>
        """

        text = classify_voc.extract_article_text(html)

        assert text is not None
        assert "회원간의 거래 분쟁에 대한 공론화 금지" not in text
        assert "가격이 얼마인가요?" in text

    def test_extract_article_text_supports_reversed_class_order(self):
        import classify_voc

        html = """
        <div class="viewer article">
          <p>사이즈가 궁금해요</p>
        </div>
        """

        text = classify_voc.extract_article_text(html)

        assert text == "사이즈가 궁금해요"

    def test_patch_sheet_xml_creates_missing_cell(self):
        import classify_voc

        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
          <sheetData>
            <row r="5"></row>
          </sheetData>
        </worksheet>
        """

        patched = classify_voc._patch_sheet_xml(xml, {"G5": "반응_문의"})

        assert b'c r="G5"' in patched
        assert b't="inlineStr"' in patched

    def test_patch_sheet_xml_preserves_ignorable_namespace_prefixes(self):
        import classify_voc

        xml = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
        <worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
          xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
          xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"
          mc:Ignorable="x14ac"
          xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac">
          <sheetFormatPr defaultRowHeight="16.5" x14ac:dyDescent="0.3"/>
          <sheetData>
            <row r="5" x14ac:dyDescent="0.3"><c r="G5" s="35" t="s"><v>25</v></c></row>
          </sheetData>
        </worksheet>
        """

        patched = classify_voc._patch_sheet_xml(xml, {"G5": "반응_문의"})
        patched_text = patched.decode("utf-8")

        assert (
            'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"'
            in patched_text
        )
        assert (
            'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"'
            in patched_text
        )
        assert 'mc:Ignorable="x14ac"' in patched_text
        assert 'x14ac:dyDescent="0.3"' in patched_text

    def test_build_output_workbook_path_uses_dated_main_prefix(self):
        import classify_voc

        output_path = classify_voc.build_output_workbook_path(
            Path("input/source.xlsx"),
            Path("output"),
            run_date=__import__("datetime").datetime(2026, 3, 25),
        )

        assert output_path == Path("output/260325_main_source.xlsx")

    def test_build_qa_artifact_workbook_path_uses_output_qa_subtree(self):
        import classify_voc

        output_path = classify_voc.build_qa_artifact_workbook_path(
            Path("input/source.xlsx"),
            artifact_name="manual-QA",
            output_dir=Path("output"),
        )

        assert output_path == Path("output/qa/manual-QA/source.xlsx")

    def test_build_qa_artifact_workbook_path_requires_artifact_name(self):
        import classify_voc

        with pytest.raises(ValueError, match="artifact_name is required"):
            classify_voc.build_qa_artifact_workbook_path(
                Path("input/source.xlsx"), artifact_name="   "
            )

    def test_build_qa_artifact_workbook_path_rejects_parent_segment(self):
        import classify_voc

        with pytest.raises(ValueError, match="single safe path segment"):
            classify_voc.build_qa_artifact_workbook_path(
                Path("input/source.xlsx"), artifact_name=".."
            )

    def test_build_qa_artifact_workbook_path_rejects_path_separator(self):
        import classify_voc

        with pytest.raises(ValueError, match="single safe path segment"):
            classify_voc.build_qa_artifact_workbook_path(
                Path("input/source.xlsx"), artifact_name="manual/qa"
            )

    @patch("classify_voc.urllib.request.urlopen")
    def test_classify_item_rule_bypasses_api(self, mock_urlopen):
        import classify_voc

        result = classify_voc.classify_item(
            "[교환] 사이즈 교환해요",
            link="https://example.com/post/1",
            api_key="test-key",
        )

        assert result == "반응_교환"
        mock_urlopen.assert_not_called()

    @patch("classify_voc.urllib.request.urlopen")
    def test_classify_item_llm_then_article_keeps_final_other(self, mock_urlopen):
        import classify_voc

        def side_effect(request, *args, **kwargs):
            if request.data is None:
                response = MagicMock()
                response.__enter__.return_value.read.return_value = (
                    '<div class="article viewer"><p>공지 외 내용 없음</p></div>'.encode(
                        "utf-8"
                    )
                )
                return response
            body = json.loads(request.data.decode())
            return _mock_http_response(
                {"choices": [{"message": {"content": "반응_기타"}}]}
            )

        mock_urlopen.side_effect = side_effect
        category, method = classify_voc._classify_item_with_method(
            "애매한 제목",
            link="https://example.com/post/1",
            api_key="test-key",
        )
        assert category == "반응_단순반응"
        assert method == "article"

    @patch("classify_voc.urllib.request.urlopen")
    def test_classify_item_llm_then_article_body(self, mock_urlopen):
        import classify_voc

        html = """
        <div class="article viewer">
          <p>회원간의 거래 분쟁에 대한 공론화 금지</p>
          <p>가격이 얼마인가요?</p>
        </div>
        """

        def side_effect(request, *args, **kwargs):
            if isinstance(request, str):
                response = MagicMock()
                response.__enter__.return_value.read.return_value = html.encode("utf-8")
                return response

            if request.data is None:
                response = MagicMock()
                response.__enter__.return_value.read.return_value = html.encode("utf-8")
                return response

            body = json.loads(request.data.decode())
            prompt = body["messages"][0]["content"]
            if "분류 대상 본문" in prompt:
                return _mock_http_response(
                    {"choices": [{"message": {"content": "반응_정보"}}]}
                )
            return _mock_http_response(
                {"choices": [{"message": {"content": "반응_기타"}}]}
            )

        mock_urlopen.side_effect = side_effect

        result = classify_voc.classify_item(
            "애매한 제목",
            link="https://example.com/post/1",
            api_key="test-key",
        )

        assert result == "반응_문의"

    @patch("classify_voc.urllib.request.urlopen")
    def test_classify_item_uses_article_when_models_fail(self, mock_urlopen):
        import classify_voc

        html = """
        <div class="article viewer">
          <p>에이치스트릿 재입고 되었어요.</p>
        </div>
        """

        def side_effect(request, *args, **kwargs):
            if getattr(request, "data", None) is None:
                response = MagicMock()
                response.__enter__.return_value.read.return_value = html.encode("cp949")
                response.__enter__.return_value.headers.get_content_charset.return_value = "ms949"
                return response
            raise _http_error(request.full_url)

        mock_urlopen.side_effect = side_effect

        category, method = classify_voc._classify_item_with_method(
            "애매한 제목",
            link="https://example.com/post/1",
            api_key="test-key",
        )

        assert category == "반응_정보"
        assert method == "article"

    def test_read_target_rows_includes_link(self):
        import classify_voc

        wb, ws = _build_base_workbook()
        ws.cell(5, 7).value = "반응_제품반응"
        ws.cell(5, 10).value = "이 제품 어디서 사나요"
        ws.cell(5, 11).value = "원문"
        ws.cell(5, 11).hyperlink = "https://example.com/post/5"

        rows = classify_voc.read_target_rows(wb)

        assert len(rows) == 1
        assert rows[0]["link"] == "https://example.com/post/5"

    @patch("classify_voc.urllib.request.urlopen")
    def test_fallback_to_gpt4o(self, mock_urlopen):
        import classify_voc

        call_count = 0

        def side_effect(request, *args, **kwargs):
            nonlocal call_count
            call_count += 1
            body = json.loads(request.data.decode())
            if body.get("model") == "openai/gpt-4.1":
                raise _http_error(request.full_url)
            if body.get("model") == "openai/gpt-4o":
                return _mock_http_response(
                    {"choices": [{"message": {"content": "반응_정보"}}]}
                )
            raise AssertionError(f"unexpected model: {body.get('model')}")

        mock_urlopen.side_effect = side_effect
        result = classify_voc.classify_title("성분이 뭔가요", api_key="test-key")
        assert result == "반응_정보"

    def test_read_target_rows(self):
        import classify_voc

        wb, ws = _build_base_workbook()
        ws.cell(5, 7).value = "반응_제품반응"
        ws.cell(5, 10).value = "이 제품 어디서 사나요"
        ws.cell(6, 7).value = "긍정_제품후기"
        ws.cell(6, 10).value = "좋아요"

        rows = classify_voc.read_target_rows(wb)

        assert len(rows) == 1
        assert rows[0]["row"] == 5
        assert rows[0]["title"] == "이 제품 어디서 사나요"
        assert all(item.get("row") != 6 for item in rows)

    def test_build_card_index_skips_legacy_prefixed_cards(self):
        import classify_voc

        wb, _ = _build_base_workbook()
        reaction = wb.create_sheet("BP_반응")
        _add_card(reaction, 2, "https://example.com/a", category="반응_제품반응")
        _add_card(reaction, 14, "https://example.com/b", category="긍정_제품후기")

        card_index = classify_voc.build_card_index(wb)

        assert card_index["https://example.com/a"] == [("BP_반응", 5, 3)]
        assert "https://example.com/b" not in card_index

    def test_write_classification_result(self):
        import classify_voc

        wb, ws = _build_base_workbook()
        ws.cell(5, 7).value = "반응_제품반응"
        ws.cell(5, 5).value = "원본값"

        classify_voc.write_classification_result(wb, row=5, category="반응_문의")

        assert ws.cell(5, 7).value == "반응_문의"
        assert ws.cell(5, 5).value == "원본값"

    def test_preserve_non_target_rows(self):
        import classify_voc

        wb, ws = _build_base_workbook()
        ws.cell(5, 7).value = "반응_제품반응"
        ws.cell(5, 5).value = "A"
        ws.cell(6, 7).value = "긍정_제품후기"
        ws.cell(6, 5).value = "B"
        ws.cell(7, 7).value = "부정_제품반응"
        ws.cell(7, 5).value = "C"

        classify_voc.write_classification_result(wb, row=5, category="반응_문의")

        assert ws.cell(6, 7).value == "긍정_제품후기"
        assert ws.cell(7, 7).value == "부정_제품반응"
        assert ws.cell(6, 5).value == "B"


@pytest.mark.integration
@patch("classify_voc.urllib.request.urlopen")
def test_classify_pipeline_with_mock_api(mock_urlopen):
    import classify_voc

    wb, ws = _build_base_workbook()
    ws.cell(5, 7).value = "반응_제품반응"
    ws.cell(5, 10).value = "재고 있나요"
    ws.cell(6, 7).value = "반응_제품반응"
    ws.cell(6, 10).value = "교환 가능한가요"
    ws.cell(7, 7).value = "반응_제품반응"
    ws.cell(7, 10).value = "가격 궁금해요"

    categories = ["반응_문의", "반응_교환", "반응_정보"]

    def side_effect(request, *args, **kwargs):
        _ = BytesIO(b"request")
        if "chat/completions" in str(getattr(request, "full_url", request)):
            payload = {"choices": [{"message": {"content": categories.pop(0)}}]}
            return _mock_http_response(payload)
        raise AssertionError(f"unexpected url: {getattr(request, 'full_url', request)}")

    mock_urlopen.side_effect = side_effect
    pipeline = _pipeline_callable(classify_voc)
    assert pipeline is not None
    pipeline(wb, api_key="test-key", output_path=None)

    assert ws.cell(5, 7).value in VALID_CATEGORIES
    assert ws.cell(6, 7).value in VALID_CATEGORIES
    assert ws.cell(7, 7).value in VALID_CATEGORIES
    assert ws.cell(5, 7).value != "반응_제품반응"
    assert ws.cell(6, 7).value != "반응_제품반응"
    assert ws.cell(7, 7).value != "반응_제품반응"


@pytest.mark.integration
@patch("classify_voc.urllib.request.urlopen")
def test_output_preserves_all_sheets(mock_urlopen):
    import classify_voc

    wb, ws = _build_base_workbook()
    ws.cell(5, 7).value = "반응_제품반응"
    ws.cell(5, 10).value = "어디서 구매"
    ws2 = wb.create_sheet("Sheet2")
    ws2.cell(1, 1).value = "keep"
    ws3 = wb.create_sheet("Sheet3")
    ws3.cell(2, 2).value = "keep-too"

    mock_urlopen.return_value = _mock_http_response(
        {"choices": [{"message": {"content": "반응_문의"}}]}
    )
    pipeline = _pipeline_callable(classify_voc)
    assert pipeline is not None
    pipeline(wb, api_key="test-key", output_path=None)

    assert wb.sheetnames == ["2026 list", "Sheet2", "Sheet3"]
    assert wb["Sheet2"].cell(1, 1).value == "keep"
    assert wb["Sheet3"].cell(2, 2).value == "keep-too"


@pytest.mark.integration
@patch("classify_voc.urllib.request.urlopen")
def test_pipeline_updates_linked_cards(mock_urlopen):
    import classify_voc

    wb, ws = _build_base_workbook()
    ws.cell(5, 7).value = "반응_제품반응"
    ws.cell(5, 10).value = "어디서 구매"
    ws.cell(5, 11).value = "원문"
    ws.cell(5, 11).hyperlink = "https://example.com/post/5"

    card_sheet = wb.create_sheet("IB")
    _add_card(card_sheet, 2, "https://example.com/post/5")
    _add_card(card_sheet, 14, "https://example.com/post/5")

    mock_urlopen.return_value = _mock_http_response(
        {"choices": [{"message": {"content": "반응_문의"}}]}
    )

    pipeline = _pipeline_callable(classify_voc)
    assert pipeline is not None
    pipeline(wb, api_key="test-key", output_path=None)

    assert ws.cell(5, 7).value == "반응_문의"
    assert card_sheet.cell(5, 3).value == "반응_문의"
    assert card_sheet.cell(5, 15).value == "반응_문의"


@pytest.mark.integration
@patch("classify_voc.urllib.request.urlopen")
def test_classification_summary(mock_urlopen, capsys):
    import classify_voc

    wb, ws = _build_base_workbook()
    ws.cell(5, 7).value = "반응_제품반응"
    ws.cell(5, 10).value = "제목1"
    ws.cell(6, 7).value = "반응_제품반응"
    ws.cell(6, 10).value = "제목2"
    ws.cell(7, 7).value = "반응_제품반응"
    ws.cell(7, 10).value = "제목3"

    mock_urlopen.return_value = _mock_http_response(
        {"choices": [{"message": {"content": "반응_문의"}}]}
    )
    pipeline = _pipeline_callable(classify_voc)
    assert pipeline is not None
    pipeline(wb, api_key="test-key", output_path=None)

    captured = capsys.readouterr()
    assert any(token in captured.err for token in ["[1/", "[2/", "[3/"])


@pytest.mark.integration
@patch("classify_voc.urllib.request.urlopen")
def test_api_error_fallback_full_pipeline(mock_urlopen):
    import classify_voc

    wb, ws = _build_base_workbook()
    ws.cell(5, 7).value = "반응_제품반응"
    ws.cell(5, 10).value = "성분 알려주세요"
    ws.cell(6, 7).value = "반응_제품반응"
    ws.cell(6, 10).value = "교환 정책은?"

    def side_effect(request, *args, **kwargs):
        body = json.loads(request.data.decode())
        if body.get("model") == "openai/gpt-4.1":
            raise _http_error(request.full_url)
        if body.get("model") == "openai/gpt-4o":
            return _mock_http_response(
                {"choices": [{"message": {"content": "반응_정보"}}]}
            )
        raise AssertionError(f"unexpected model: {body.get('model')}")

    mock_urlopen.side_effect = side_effect
    pipeline = _pipeline_callable(classify_voc)
    assert pipeline is not None
    pipeline(wb, api_key="test-key", output_path=None)

    assert ws.cell(5, 7).value in VALID_CATEGORIES
    assert ws.cell(6, 7).value in VALID_CATEGORIES
    assert ws.cell(5, 7).value != "반응_제품반응"
    assert ws.cell(6, 7).value != "반응_제품반응"


@pytest.mark.integration
@patch("classify_voc.urllib.request.urlopen")
def test_hybrid_pipeline_integration(mock_urlopen):
    import classify_voc

    wb, ws = _build_base_workbook()
    ws.cell(5, 7).value = "반응_제품반응"
    ws.cell(5, 10).value = "[교환] 140으로 바꿔요"
    ws.cell(6, 7).value = "반응_제품반응"
    ws.cell(6, 10).value = "가격 궁금합니다"

    mock_urlopen.return_value = _mock_http_response(
        {"choices": [{"message": {"content": "반응_정보"}}]}
    )
    pipeline = _pipeline_callable(classify_voc)
    assert pipeline is not None
    pipeline(wb, api_key="test-key", output_path=None)

    assert ws.cell(5, 7).value == "반응_교환"
    assert ws.cell(6, 7).value == "반응_문의"
    assert mock_urlopen.call_count == 0


@pytest.mark.integration
def test_openpyxl_roundtrip_drops_media_on_real_workbook():
    src = _real_input_workbook()
    before_media, before_drawings = _xlsx_part_counts(src)

    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / "roundtrip.xlsx"
        wb = openpyxl.load_workbook(src)
        wb.save(out)

        after_media, after_drawings = _xlsx_part_counts(out)

    assert after_media < before_media
    assert after_drawings < before_drawings


@pytest.mark.integration
@patch("classify_voc.urllib.request.urlopen")
def test_saved_output_preserves_media_on_real_workbook(mock_urlopen):
    import classify_voc

    src = _real_input_workbook()
    before_media, before_drawings = _xlsx_part_counts(src)
    out = Path(tempfile.mkdtemp()) / src.name

    mock_urlopen.return_value = _mock_http_response(
        {"choices": [{"message": {"content": "반응_문의"}}]}
    )

    wb = openpyxl.load_workbook(src)
    classify_voc.run_classification(
        wb, api_key="test-key", output_path=out, source_path=src
    )

    after_media, after_drawings = _xlsx_part_counts(out)

    assert after_media == before_media
    assert after_drawings == before_drawings


@pytest.mark.integration
def test_saved_output_reloads_supported_workbook_updates():
    import classify_voc

    src = _real_input_workbook()
    updates = _build_legacy_baseline_updates(src)
    expected_updates = _updates_by_sheet(updates)

    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / "supported-result.xlsx"
        classify_voc.save_workbook_with_preserved_media(src, out, updates)

        saved = openpyxl.load_workbook(out)

    for sheet_name, cells in expected_updates.items():
        ws = saved[sheet_name]
        for cell_ref, expected_value in cells.items():
            assert ws[cell_ref].value == expected_value


@pytest.mark.integration
@patch("classify_voc.urllib.request.urlopen")
def test_saved_output_synthetic_workbook_with_linked_cards_hard_fails(mock_urlopen):
    import classify_voc

    with tempfile.TemporaryDirectory() as tmpdir:
        src = Path(tmpdir) / "source.xlsx"
        out = Path(tmpdir) / "result.xlsx"

        wb, ws = _build_base_workbook()
        ws.cell(5, 7).value = "반응_제품반응"
        ws.cell(5, 10).value = "어디서 구매"
        ws.cell(5, 11).value = "원문"
        ws.cell(5, 11).hyperlink = "https://example.com/post/5"
        card_sheet = wb.create_sheet("IB")
        _add_card(card_sheet, 2, "https://example.com/post/5")
        bp_sheet = wb.create_sheet("BP_긍정")
        _add_card(bp_sheet, 20, "https://example.com/post/5")
        wb.save(src)

        mock_urlopen.return_value = _mock_http_response(
            {"choices": [{"message": {"content": "반응_문의"}}]}
        )

        loaded = openpyxl.load_workbook(src)
        with pytest.raises(classify_voc.UnsupportedWorkbookVariantError) as excinfo:
            classify_voc.run_classification(
                loaded, api_key="test-key", output_path=out, source_path=src
            )

    assert excinfo.value.policy == "hard_fail"
    assert excinfo.value.reason_code == "unsupported_sheet_count"
    assert not out.exists()


@pytest.mark.integration
@patch("classify_voc.urllib.request.urlopen")
def test_saved_output_synthetic_workbook_without_linked_card_mapping_hard_fails(
    mock_urlopen,
):
    import classify_voc

    with tempfile.TemporaryDirectory() as tmpdir:
        src = Path(tmpdir) / "source.xlsx"
        out = Path(tmpdir) / "result.xlsx"

        wb, ws = _build_base_workbook()
        ws.cell(5, 7).value = "반응_제품반응"
        ws.cell(5, 10).value = "가격 궁금해요"
        wb.save(src)

        mock_urlopen.return_value = _mock_http_response(
            {"choices": [{"message": {"content": "반응_정보"}}]}
        )

        loaded = openpyxl.load_workbook(src)
        with pytest.raises(classify_voc.UnsupportedWorkbookVariantError) as excinfo:
            classify_voc.run_classification(
                loaded, api_key="test-key", output_path=out, source_path=src
            )

    assert excinfo.value.policy == "hard_fail"
    assert excinfo.value.reason_code == "unsupported_sheet_count"
    assert not out.exists()


@pytest.mark.integration
@patch("classify_voc.urllib.request.urlopen")
def test_saved_output_preserves_worksheet_namespace_prefixes(mock_urlopen):
    import classify_voc

    src = _real_input_workbook()
    out = Path(tempfile.mkdtemp()) / src.name

    mock_urlopen.return_value = _mock_http_response(
        {"choices": [{"message": {"content": "반응_문의"}}]}
    )

    wb = openpyxl.load_workbook(src)
    classify_voc.run_classification(
        wb, api_key="test-key", output_path=out, source_path=src
    )

    with ZipFile(out) as zf:
        sheet_xml = zf.read("xl/worksheets/sheet2.xml").decode("utf-8")

    xmlns = dict(re.findall(r'xmlns(?::([^=]+))?="([^"]+)"', sheet_xml[:600]))
    ignorable_match = re.search(r'([A-Za-z0-9_]+):Ignorable="([^"]+)"', sheet_xml[:600])

    assert xmlns["mc"] == "http://schemas.openxmlformats.org/markup-compatibility/2006"
    assert (
        xmlns["x14ac"] == "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
    )
    assert ignorable_match is not None
    assert ignorable_match.group(1) == "mc"
    assert ignorable_match.group(2) == "x14ac"


def test_apply_layout_to_output_workbook_formats_pages():
    import classify_voc

    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "layout.xlsx"
        wb, _ = _build_base_workbook()
        wb.create_sheet("주현황보고", 0)
        ws = wb.create_sheet("IB")
        ws.cell(2, 2).value = "Voice Of Customer"
        ws.cell(2, 8).value = "Voice Of Customer"
        ws.cell(5, 3).value = "반응_제품반응"
        ws.cell(5, 9).value = "반응_제품반응"
        _apply_card_outline(ws, 2)
        _apply_card_outline(ws, 8)
        original_left_margin = ws.page_margins.left
        ws.print_area = "'IB'!$B$2:$M$35"
        wb.save(path)

        classify_voc._apply_layout_to_output_workbook(path)

        saved = openpyxl.load_workbook(path)
        with ZipFile(path) as zf:
            workbook_xml = zf.read("xl/workbook.xml").decode("utf-8")

    assert "주현황보고" in saved.sheetnames
    assert saved["IB"]["B2"].value == "Voice Of Customer"
    assert saved["IB"]["I2"].value == "Voice Of Customer"
    assert saved["IB"]["H2"].value is None
    assert saved["IB"]["O2"].value is None
    assert saved["IB"].print_area == "'IB'!$B$2:$O$35"
    assert saved["IB"].column_dimensions["H"].width == pytest.approx(0.875)
    assert saved["IB"].column_dimensions["O"].width == pytest.approx(0.875)
    assert saved["IB"]["G2"].border.right.style == "medium"
    assert saved["IB"]["N2"].border.right.style == "medium"
    assert saved["IB"].print_options.horizontalCentered is True
    assert saved["IB"].print_options.verticalCentered is True
    assert saved["IB"].page_setup.pageOrder == "overThenDown"
    assert saved["IB"].page_margins.left == pytest.approx(original_left_margin + 0.2)
    assert 'name="주현황보고"' in workbook_xml


@pytest.mark.integration
@patch("classify_voc.urllib.request.urlopen")
def test_primary_template_value_only_output_passes_package_validator(mock_urlopen):
    import classify_voc

    src = _real_input_workbook()
    out = Path(tempfile.mkdtemp()) / src.name

    mock_urlopen.return_value = _mock_http_response(
        {"choices": [{"message": {"content": "반응_문의"}}]}
    )

    wb = openpyxl.load_workbook(src)
    classify_voc.run_classification(
        wb, api_key="test-key", output_path=out, source_path=src
    )

    result = subprocess.run(
        _xlsx_package_diff_validator_cmd(src, out),
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0, result.stdout + result.stderr
    assert "PASS:" in result.stdout


@pytest.mark.integration
def test_xlsx_baseline_matrix_real_workbook():
    import classify_voc

    src = _real_input_workbook()
    updates = _build_legacy_baseline_updates(src)

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        value_only = tmp / "value-only.xlsx"
        layout_only = tmp / "layout-only.xlsx"
        mixed = tmp / "mixed-legacy.xlsx"

        _save_value_only_patch(src, value_only, updates)
        shutil.copy2(src, layout_only)
        classify_voc._apply_layout_to_output_workbook(layout_only)
        classify_voc.save_workbook_with_preserved_media_legacy(src, mixed, updates)

        source_cols = _sheet_max_col_map(src)
        value_only_cols = _sheet_max_col_map(value_only)
        layout_only_cols = _sheet_max_col_map(layout_only)
        mixed_cols = _sheet_max_col_map(mixed)

        source_parts = _xlsx_part_counts(src)
        value_only_parts = _xlsx_part_counts(value_only)
        layout_only_parts = _xlsx_part_counts(layout_only)
        mixed_parts = _xlsx_part_counts(mixed)

        value_only_changed = _changed_zip_parts(src, value_only)
        layout_only_changed = _changed_zip_parts(src, layout_only)
        mixed_changed = _changed_zip_parts(src, mixed)

    assert value_only_parts == source_parts
    assert layout_only_parts == source_parts
    assert mixed_parts == source_parts

    assert value_only_cols["2026 list"] == source_cols["2026 list"]
    assert any(
        layout_only_cols[name] > source_cols[name]
        for name in source_cols
        if name != "2026 list" and name in layout_only_cols
    )
    assert any(
        mixed_cols[name] > source_cols[name]
        for name in source_cols
        if name != "2026 list" and name in mixed_cols
    )

    assert len(value_only_changed) < len(mixed_changed)
    assert "xl/workbook.xml" not in value_only_changed
    assert "xl/workbook.xml" in layout_only_changed
    assert "xl/workbook.xml" in mixed_changed


@pytest.mark.integration
def test_workbook_topology_preserved_baseline_legacy_writer_not_preserved():
    import classify_voc

    src = _real_input_workbook()
    updates = _build_legacy_baseline_updates(src)

    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / "legacy-output.xlsx"
        classify_voc.save_workbook_with_preserved_media_legacy(src, out, updates)

        source_cols = _sheet_max_col_map(src)
        output_cols = _sheet_max_col_map(out)

    increased = [
        name
        for name, src_col in source_cols.items()
        if name != "2026 list" and output_cols.get(name, src_col) > src_col
    ]
    assert increased


@pytest.mark.integration
def test_package_diff_allowlist_baseline_legacy_writer_has_wide_changes():
    import classify_voc

    src = _real_input_workbook()
    updates = _build_legacy_baseline_updates(src)

    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / "legacy-output.xlsx"
        classify_voc.save_workbook_with_preserved_media_legacy(src, out, updates)
        changed = _changed_zip_parts(src, out)

        with ZipFile(src) as zf:
            sheet_paths = classify_voc._sheet_xml_paths(zf)
        all_sheet_xml = set(sheet_paths.values())

    changed_sheet_xml = {name for name in changed if name in all_sheet_xml}
    changed_drawings = {
        name
        for name in changed
        if name.startswith("xl/drawings/") and name.endswith(".xml")
    }

    assert "xl/workbook.xml" in changed
    assert changed_sheet_xml
    assert changed_drawings


@pytest.mark.integration
def test_writable_surface_contract_supported_template():
    xlsx_template_contract = importlib.import_module("xlsx_template_contract")

    src = _real_input_workbook()
    contract = _supported_workbook_contract()

    fingerprint = xlsx_template_contract.validate_supported_workbook(src, contract)
    allowed = xlsx_template_contract.expand_allowed_writable_surface(contract)

    assert fingerprint["workbook_xml_sha256"] in _contract_workbook_hashes(contract)
    assert fingerprint["sheet_order"] == contract["preserved_topology"]["sheet_order"]
    assert fingerprint["sheet_states"] in _contract_sheet_state_variants(contract)
    assert fingerprint["linked_card_targets"] in _contract_linked_card_target_variants(
        contract
    )
    assert len(allowed["2026 list"]) == 216
    assert (
        sum(
            len(cell_refs)
            for sheet_name, cell_refs in allowed.items()
            if sheet_name != "2026 list"
        )
        == 162
    )


@pytest.mark.integration
def test_writable_surface_contract_supports_family_without_file_hash_lock():
    xlsx_template_contract = importlib.import_module("xlsx_template_contract")

    src = _real_input_workbook()
    contract = copy.deepcopy(_supported_workbook_contract())
    contract["supported_workbook_family"]["file_sha256"] = "0" * 64

    fingerprint = xlsx_template_contract.validate_supported_workbook(src, contract)

    assert fingerprint["sheet_order"] == contract["preserved_topology"]["sheet_order"]
    assert (
        fingerprint["linked_card_targets"]
        == contract["writable_surface"]["linked_card_targets"]
    )


@pytest.mark.integration
def test_writable_surface_contract_rejects_sheet_state_drift():
    xlsx_template_contract = importlib.import_module("xlsx_template_contract")

    src = _real_input_workbook()
    contract = copy.deepcopy(_supported_workbook_contract())
    for variant in contract["preserved_topology"].get("sheet_states_variants", []):
        variant["주현황보고"] = "visible"
    if "sheet_states" in contract["preserved_topology"]:
        contract["preserved_topology"]["sheet_states"]["주현황보고"] = "visible"

    with pytest.raises(xlsx_template_contract.WorkbookContractError) as excinfo:
        xlsx_template_contract.validate_supported_workbook(src, contract)

    assert excinfo.value.code == "unsupported_sheet_state"


@pytest.mark.integration
def test_writable_surface_contract_allows_supported_updates():
    xlsx_template_contract = importlib.import_module("xlsx_template_contract")

    src = _real_input_workbook()
    contract = _supported_workbook_contract()
    updates = _build_legacy_baseline_updates(src)
    fingerprint = xlsx_template_contract.validate_supported_workbook(src, contract)
    resolved_contract = xlsx_template_contract.resolve_runtime_contract(
        contract, fingerprint
    )

    xlsx_template_contract.enforce_writable_surface(updates, resolved_contract)


def test_writable_surface_rejects_unknown_cell():
    xlsx_template_contract = importlib.import_module("xlsx_template_contract")

    contract = _supported_workbook_contract()
    updates = [
        {
            "sheet_name": "2026 list",
            "cell_ref": "H5",
            "value": "반응_문의",
        }
    ]

    with pytest.raises(xlsx_template_contract.WorkbookContractError) as excinfo:
        xlsx_template_contract.enforce_writable_surface(updates, contract)

    assert excinfo.value.code == "unsupported_write_cell"
    assert "2026 list!H5" in str(excinfo.value)


@pytest.mark.integration
def test_primary_writer_rejects_structural_edit():
    import classify_voc

    src = _real_input_workbook()
    updates: list[classify_voc.CellUpdate] = [
        {
            "sheet_name": "2026 list",
            "cell_ref": "H5",
            "value": "반응_문의",
        }
    ]

    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / "structural-edit.xlsx"
        with pytest.raises(classify_voc.UnsupportedWorkbookVariantError) as excinfo:
            classify_voc.save_workbook_with_preserved_media(src, out, updates)

    assert excinfo.value.reason_code == "unsupported_write_cell"
    assert excinfo.value.policy == "hard_fail"


@pytest.mark.integration
@patch("classify_voc.urllib.request.urlopen")
def test_run_classification_primary_path_skips_layout_mutation(mock_urlopen):
    import classify_voc

    src = _real_input_workbook()
    out = Path(tempfile.mkdtemp()) / src.name

    mock_urlopen.return_value = _mock_http_response(
        {"choices": [{"message": {"content": "반응_문의"}}]}
    )

    wb = openpyxl.load_workbook(src)
    with patch.object(
        classify_voc,
        "_apply_output_layout_to_workbook",
        side_effect=AssertionError(
            "layout mutation path should not run in primary writer"
        ),
    ):
        classify_voc.run_classification(
            wb, api_key="test-key", output_path=out, source_path=src
        )


@pytest.mark.integration
def test_supported_variant_uses_primary_writer(capsys):
    import classify_voc

    src = _real_input_workbook()
    updates = _build_legacy_baseline_updates(src)

    selection = classify_voc.select_output_writer(src, updates)
    assert selection["supported"] is True
    assert selection["writer"] == "primary_template_value_only"
    assert selection["policy"] == "hard_fail"
    assert selection["reason_code"] is None

    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / src.name
        classify_voc.save_workbook_with_preserved_media(src, out, updates)
        assert out.exists()

    captured = capsys.readouterr()
    assert "output_writer=primary_template_value_only policy=hard_fail" in captured.err


@pytest.mark.integration
def test_unsupported_variant_detection_hard_fail_reason_code(capsys):
    import classify_voc

    src = _real_input_workbook()
    updates = _build_legacy_baseline_updates(src)

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        unsupported = tmpdir_path / "unsupported-sheet-state.xlsx"
        out = tmpdir_path / "output.xlsx"
        shutil.copy2(src, unsupported)
        _rewrite_zip_member(
            unsupported,
            "xl/workbook.xml",
            lambda data: data.replace(b'state="hidden"', b'state="visible"', 1),
        )

        selection = classify_voc.select_output_writer(unsupported, updates)
        assert selection["supported"] is False
        assert selection["writer"] == "unsupported_variant"
        assert selection["policy"] == "hard_fail"
        assert selection["reason_code"] == "unsupported_sheet_state"

        with pytest.raises(classify_voc.UnsupportedWorkbookVariantError) as excinfo:
            classify_voc.save_workbook_with_preserved_media(unsupported, out, updates)

        assert excinfo.value.reason_code == "unsupported_sheet_state"
        assert excinfo.value.policy == "hard_fail"
        assert not out.exists()

    captured = capsys.readouterr()
    assert (
        "output_writer=unsupported policy=hard_fail reason_code=unsupported_sheet_state"
        in captured.err
    )


@pytest.mark.integration
def test_workbook_topology_preserved_for_allowlisted_value_only_output():
    src = _real_input_workbook()
    updates = _build_legacy_baseline_updates(src)

    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / "value-only.xlsx"
        _save_value_only_patch(src, out, updates)
        result = subprocess.run(
            _xlsx_package_diff_validator_cmd(src, out),
            capture_output=True,
            text=True,
            check=False,
        )

    assert result.returncode == 0, result.stdout + result.stderr
    assert "PASS:" in result.stdout


@pytest.mark.integration
def test_package_diff_allowlist_rejects_unexpected_change():
    src = _real_input_workbook()
    updates = _build_legacy_baseline_updates(src)

    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / "value-only-mutated.xlsx"
        _save_value_only_patch(src, out, updates)
        _rewrite_zip_member(
            out, "xl/_rels/workbook.xml.rels", lambda data: data + b"\n"
        )
        result = subprocess.run(
            _xlsx_package_diff_validator_cmd(src, out),
            capture_output=True,
            text=True,
            check=False,
        )

    assert result.returncode == 1, result.stdout + result.stderr
    assert "unexpected changed package parts" in result.stdout
    assert "xl/_rels/workbook.xml.rels" in result.stdout


@pytest.mark.integration
def test_package_diff_allowlist_rejects_non_writable_cell_drift_inside_writable_sheet_xml():
    import classify_voc

    xlsx_template_contract = importlib.import_module("xlsx_template_contract")

    src = _real_input_workbook()
    updates = _build_legacy_baseline_updates(src)
    contract = _supported_workbook_contract()
    writable_surface = xlsx_template_contract.expand_allowed_writable_surface(contract)

    with ZipFile(src) as zf:
        sheet_path = classify_voc._sheet_xml_paths(zf)[
            contract["writable_surface"]["primary_target"]["sheet_name"]
        ]

    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / "value-only-writable-drift.xlsx"
        _save_value_only_patch(src, out, updates)
        _rewrite_zip_member(
            out,
            sheet_path,
            lambda data: _mutate_first_non_writable_cell(
                data,
                set(
                    writable_surface[
                        contract["writable_surface"]["primary_target"]["sheet_name"]
                    ]
                ),
            ),
        )
        result = subprocess.run(
            _xlsx_package_diff_validator_cmd(src, out),
            capture_output=True,
            text=True,
            check=False,
        )

    assert result.returncode == 1, result.stdout + result.stderr
    assert (
        "allowlisted writable sheet contains structural/layout drift" in result.stdout
    )
