"""VOC classification script using GitHub Models API."""

import argparse
import html
import importlib
import json
import math
import os
import re
import sys
import time
import urllib.request
import urllib.error
import xml.etree.ElementTree as ET
from pathlib import Path
from posixpath import dirname, normpath, join as posix_join
from zipfile import ZipFile
from typing import Any, Mapping, Sequence, TypedDict, cast

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import (
    column_index_from_string,
    quote_sheetname,
    range_boundaries,
)

VALID_CATEGORIES = {
    "반응_문의",
    "반응_교환",
    "반응_구해요",
    "반응_정보",
    "반응_기타",
}
TARGET_SHEET = "2026 list"
TARGET_MARKER = "반응_제품반응"
TARGET_COLUMN = 7
TITLE_COLUMN = 10
LINK_COLUMN = 11
START_ROW = 5
END_ROW = 220

EXCHANGE_KEYWORDS = [
    "[교환]",
    "교환원해",
    "교환 구해",
    "교환하실",
    "교환해주",
    "교환 원해",
]
SEEKING_KEYWORDS = [
    "[구해요]",
    "구해요",
    "구해용",
    "구합니다",
    "구입하실분",
    "구해봐",
    "구해보아요",
    "구매원해",
    "구매합니다",
    "구함",
]
INFO_KEYWORDS = [
    "가격",
    "무배",
    "할인",
    "세일",
    "핫딜",
    "리셀",
    "발매",
    "신상",
]

NOTICE_PHRASES = ("회원간의 거래 분쟁에 대한 공론화 금지",)
XML_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XML_NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
XML_NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
XML_NS_MC = "http://schemas.openxmlformats.org/markup-compatibility/2006"
XML_NS_X14AC = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
XML_NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
DELETE_SHEET_NAME = "주현황보고"
LAYOUT_EXCLUDED_SHEET = TARGET_SHEET
PAGE_START_COLUMN = 2
PAGE_CONTENT_WIDTH = 6
PAGE_TOTAL_WIDTH = 7
PAGE_START_ROW = 2
PAGE_END_ROW = 35
PAD_COLUMN_WIDTH = 0.875
THICK_BLACK_SIDE = Side(style="medium", color="000000")
PRIMARY_WRITER_ID = "primary_template_value_only"
UNSUPPORTED_VARIANT_POLICY = "hard_fail"

ET.register_namespace("", XML_NS_MAIN)
ET.register_namespace("r", XML_NS_REL)
ET.register_namespace("mc", XML_NS_MC)
ET.register_namespace("x14ac", XML_NS_X14AC)
ET.register_namespace("xdr", XML_NS_XDR)


class TargetRow(TypedDict):
    row: int
    title: str
    link: str


class CellUpdate(TypedDict):
    sheet_name: str
    cell_ref: str
    value: str


class OutputWriterSelection(TypedDict):
    writer: str
    policy: str
    supported: bool
    reason_code: str | None
    detail: str | None


class UnsupportedWorkbookVariantError(RuntimeError):
    reason_code: str
    detail: str
    policy: str

    def __init__(self, reason_code: str, detail: str):
        self.reason_code = reason_code
        self.detail = detail
        self.policy = UNSUPPORTED_VARIANT_POLICY
        super().__init__(f"{self.policy}:{reason_code}: {detail}")


def preprocess_title(text) -> str | None:
    if text is None:
        return None
    cleaned = str(text).replace("\n", " ").strip()
    if not cleaned:
        return None
    return " ".join(cleaned.split())


def build_classification_prompt(title) -> str:
    categories = "\n".join(f"- {item}" for item in sorted(VALID_CATEGORIES))
    return (
        "당신은 VOC 제목을 아래 5개 카테고리 중 하나로만 분류하는 분류기입니다.\n"
        "카테고리 목록:\n"
        f"{categories}\n\n"
        "카테고리 분류 기준:\n"
        "- 반응_교환: [교환] 태그, 교환 요청, 사이즈 교환, 교환 구해요\n"
        "- 반응_구해요: [구해요] 태그, 구해요, 구합니다, 구해용, 구입하실분, 판매처 문의\n"
        "- 반응_문의: 직접 답변이나 조언을 요청하는 질문형 제목. 궁금, 질문, 어디서 사나요, ~인가요?, 있을까요?, 계신가요?\n"
        "- 반응_정보: 가격, 성분, 할인, 세일, 무배, 핫딜, 리셀, 발매, 신상, 구매 정보나 경험 정보를 공유하는 제목\n"
        "- 반응_기타: 위 4개에 해당하지 않는 단순 감상, 후기, 추천, 비추천\n\n"
        "few-shot 예시:\n"
        '입력: "궁금합니다" -> 출력: 반응_문의\n'
        '입력: "어디서 사나요" -> 출력: 반응_문의\n'
        '입력: "구매처 아시는분" -> 출력: 반응_문의\n'
        '입력: "판매처 아시는분" -> 출력: 반응_구해요\n'
        '입력: "불량 수선받으신 맘님 계신가요" -> 출력: 반응_문의\n'
        '입력: "[교환] 교환해주세요" -> 출력: 반응_교환\n'
        '입력: "스피드캣 교환원해요" -> 출력: 반응_교환\n'
        '입력: "[교환] 블랙150(저) > 140(맘님)" -> 출력: 반응_교환\n'
        '입력: "구합니다" -> 출력: 반응_구해요\n'
        '입력: "구해요" -> 출력: 반응_구해요\n'
        '입력: "[구해요] 실버 130 구해용" -> 출력: 반응_구해요\n'
        '입력: "구입하실분들" -> 출력: 반응_구해요\n'
        '입력: "가격이 얼마인가요" -> 출력: 반응_정보\n'
        '입력: "성분 정보" -> 출력: 반응_정보\n'
        '입력: "무신사 79000원 무배" -> 출력: 반응_정보\n'
        '입력: "크림 리셀가 공유" -> 출력: 반응_정보\n'
        '입력: "최대 80% 할인 핫딜" -> 출력: 반응_정보\n'
        '입력: "신상 발매 정보" -> 출력: 반응_정보\n'
        '입력: "왕비추천" -> 출력: 반응_기타\n'
        '입력: "베베드피노 몬치치가방" -> 출력: 반응_기타\n\n'
        "중요: [교환] 또는 [구해요] 태그가 있으면 해당 카테고리로 분류하세요.\n"
        "중요: 가격/세일/무배/리셀/핫딜/발매/신상처럼 정보를 공유하는 제목은 질문형 어미가 섞여도 반응_정보를 우선 검토하세요.\n"
        "반드시 카테고리 텍스트 하나만 출력하세요.\n"
        f"분류 대상 제목: {title}"
    )


def build_body_classification_prompt(text) -> str:
    categories = "\n".join(f"- {item}" for item in sorted(VALID_CATEGORIES))
    return (
        "당신은 VOC 본문을 아래 5개 카테고리 중 하나로만 분류하는 분류기입니다.\n"
        "카테고리 목록:\n"
        f"{categories}\n\n"
        "카테고리 기준:\n"
        "- 반응_교환: 교환 요청, 사이즈 교환, 맞교환\n"
        "- 반응_구해요: 구매 희망, 구합니다, 구해요, 판매처 수배\n"
        "- 반응_문의: 직접 답변이나 조언을 구하는 질문\n"
        "- 반응_정보: 가격, 할인, 무배, 리셀, 발매, 신상, 경험/상황 공유\n"
        "- 반응_기타: 감상, 잡담, 단순 반응\n"
        "본문에는 공지와 배너가 제거되어 있을 수 있습니다. 실제 게시글 본문 의미만 보고 분류하세요.\n"
        "반드시 카테고리 텍스트 하나만 출력하세요.\n"
        f"분류 대상 본문: {text}"
    )


def validate_result(text) -> str:
    value = (text or "").strip()
    if value in VALID_CATEGORIES:
        return value
    for cat in VALID_CATEGORIES:
        if cat in value:
            return cat
    return "반응_기타"


def _request_json(url: str, body: dict[str, Any], api_key: str) -> dict[str, Any]:
    data = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
            "User-Agent": "voc-classifier/1.0",
        },
        method="POST",
    )
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read().decode("utf-8"))


class GithubModelsAdapter:
    endpoint = "https://models.github.ai/inference/chat/completions"

    def __init__(
        self, api_key: str, model: str = "openai/gpt-4.1", prompt_builder=None
    ):
        self.api_key = api_key
        self.model = model
        self.prompt_builder = prompt_builder or build_classification_prompt

    def build_request_body(self, title):
        return {
            "model": self.model,
            "messages": [
                {
                    "role": "system",
                    "content": self.prompt_builder(title),
                },
                {
                    "role": "user",
                    "content": str(title),
                },
            ],
        }

    def parse_response(self, resp):
        return resp["choices"][0]["message"]["content"]


def _classify_with_adapter(adapter, title) -> str | None:
    """Returns None on API failure (HTTP/network), triggering fallback."""
    backoff = 1
    for attempt in range(3):
        try:
            body = adapter.build_request_body(title)
            response = _request_json(adapter.endpoint, body, adapter.api_key)
            return validate_result(adapter.parse_response(response))
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError):
            if attempt == 2:
                return None
            time.sleep(backoff)
            backoff *= 2
        except (KeyError, IndexError, TypeError, ValueError):
            return "반응_기타"
        finally:
            time.sleep(0.5)
    return None


def rule_classify(title: str) -> str | None:
    lower = title.lower()
    for kw in EXCHANGE_KEYWORDS:
        if kw.lower() in lower:
            return "반응_교환"
    for kw in SEEKING_KEYWORDS:
        if kw.lower() in lower:
            return "반응_구해요"
    if re.search(r"\b\d[\d,]*원\b", title):
        return "반응_정보"
    for kw in INFO_KEYWORDS:
        if kw.lower() in lower:
            return "반응_정보"
    return None


def _extract_link_text(cell) -> str:
    if cell.hyperlink and getattr(cell.hyperlink, "target", None):
        return str(cell.hyperlink.target)
    return str(cell.value or "")


def _fetch_text(url: str) -> str | None:
    if not url:
        return None
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "voc-classifier/1.0",
        },
        method="GET",
    )
    try:
        with urllib.request.urlopen(req) as resp:
            data = resp.read()
    except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError):
        return None

    if not data:
        return None
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError:
        return data.decode("utf-8", errors="ignore")


def extract_article_text(html_text: str) -> str | None:
    if not html_text:
        return None

    container_html = html_text
    selectors = (
        r"<(?P<tag>[a-zA-Z0-9]+)[^>]*class=[\"'][^\"']*\barticle\b[^\"']*\bviewer\b[^\"']*[\"'][^>]*>(?P<content>.*?)</(?P=tag)>",
        r"<(?P<tag>[a-zA-Z0-9]+)[^>]*class=[\"'][^\"']*\bviewer\b[^\"']*\barticle\b[^\"']*[\"'][^>]*>(?P<content>.*?)</(?P=tag)>",
        r"<(?P<tag>[a-zA-Z0-9]+)[^>]*class=[\"'][^\"']*\bartice\b[^\"']*\bviewer\b[^\"']*[\"'][^>]*>(?P<content>.*?)</(?P=tag)>",
        r"<(?P<tag>[a-zA-Z0-9]+)[^>]*class=[\"'][^\"']*\bviewer\b[^\"']*\bartice\b[^\"']*[\"'][^>]*>(?P<content>.*?)</(?P=tag)>",
    )
    for pattern in selectors:
        match = re.search(pattern, html_text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            container_html = match.group("content")
            break

    container_html = re.sub(
        r"<(script|style|noscript)\b[^>]*>.*?</\1>",
        " ",
        container_html,
        flags=re.IGNORECASE | re.DOTALL,
    )

    for phrase in NOTICE_PHRASES:
        container_html = re.sub(
            rf"<[^>]+>[^<]*{re.escape(phrase)}[^<]*</[^>]+>",
            " ",
            container_html,
            flags=re.IGNORECASE | re.DOTALL,
        )
        container_html = container_html.replace(phrase, " ")

    text = re.sub(r"<[^>]+>", "\n", container_html)
    text = html.unescape(text)
    text = re.sub(r"\n+", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    cleaned = text.strip()
    return cleaned or None


def classify_article_text(text: str, api_key: str) -> str | None:
    primary_result = _classify_with_adapter(
        GithubModelsAdapter(
            api_key=api_key,
            model="openai/gpt-4.1",
            prompt_builder=build_body_classification_prompt,
        ),
        text,
    )
    if primary_result is not None:
        return primary_result
    return _classify_with_adapter(
        GithubModelsAdapter(
            api_key=api_key,
            model="openai/gpt-4o",
            prompt_builder=build_body_classification_prompt,
        ),
        text,
    )


def classify_from_article(link: str | None, api_key: str) -> str | None:
    if not link:
        return None
    html = _fetch_text(link)
    if html is None:
        return None
    text = extract_article_text(html)
    if text is None:
        return None
    return classify_article_text(text, api_key)


def build_card_index(wb) -> dict[str, list[tuple[str, int, int]]]:
    index: dict[str, list[tuple[str, int, int]]] = {}
    for ws in wb.worksheets:
        card_starts: list[int] = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(2, col).value
            if value == "Voice Of Customer":
                card_starts.append(col)

        for card_start in card_starts:
            link = _extract_link_text(ws.cell(4, card_start + 4))
            if not link:
                continue
            if ws.cell(5, card_start).value != "주요 이슈":
                continue
            index.setdefault(link, []).append((ws.title, 5, card_start + 1))
    return index


def write_linked_classification_results(
    wb,
    link: str | None,
    category: str,
    card_index: dict[str, list[tuple[str, int, int]]] | None,
):
    if not link or not card_index:
        return

    for sheet_name, row, col in card_index.get(link, []):
        ws = wb[sheet_name]
        ws.cell(row, col).value = validate_result(category)


def _classify_item_with_method(
    title,
    link: str | None,
    api_key,
) -> tuple[str, str]:
    cleaned = preprocess_title(title)
    if cleaned is None:
        return "반응_기타", "empty"

    rule_result = rule_classify(cleaned)
    if rule_result is not None:
        return rule_result, "rule"

    primary_result = _classify_with_adapter(
        GithubModelsAdapter(api_key=api_key, model="openai/gpt-4.1"), cleaned
    )
    if primary_result is not None and primary_result != "반응_기타":
        return primary_result, "llm"

    if primary_result == "반응_기타" and link:
        article_result = classify_from_article(link, api_key)
        if article_result is not None and article_result != "반응_기타":
            return article_result, "article"

    if primary_result == "반응_기타":
        return "반응_기타", "article"

    fallback_result = _classify_with_adapter(
        GithubModelsAdapter(api_key=api_key, model="openai/gpt-4o"), cleaned
    )
    if fallback_result is not None:
        return fallback_result, "fallback"
    return "반응_기타", "fallback"


def classify_item(
    title,
    link: str | None,
    api_key,
) -> str:
    category, _ = _classify_item_with_method(title, link, api_key)
    return category


def classify_title(title, api_key) -> str:
    return classify_item(title, link=None, api_key=api_key)


def read_target_rows(wb) -> list[TargetRow]:
    ws = wb[TARGET_SHEET]
    targets = []
    for row in range(START_ROW, END_ROW + 1):
        if ws.cell(row, TARGET_COLUMN).value != TARGET_MARKER:
            continue
        title = preprocess_title(ws.cell(row, TITLE_COLUMN).value)
        if title is None:
            continue
        cell = ws.cell(row, LINK_COLUMN)
        link_val = _extract_link_text(cell)
        targets.append({"row": row, "title": title, "link": link_val})
    return targets


def write_classification_result(wb, row, category):
    ws = wb[TARGET_SHEET]
    ws.cell(row, TARGET_COLUMN).value = validate_result(category)


def _cell_ref(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def _split_cell_ref(cell_ref: str) -> tuple[int, str]:
    match = re.fullmatch(r"([A-Z]+)(\d+)", cell_ref)
    if match is None:
        raise ValueError(f"invalid cell reference: {cell_ref}")
    return int(match.group(2)), match.group(1)


def _build_sheet_updates(
    row: int,
    link: str | None,
    category: str,
    card_index: dict[str, list[tuple[str, int, int]]] | None,
) -> list[CellUpdate]:
    updates: list[CellUpdate] = [
        {
            "sheet_name": TARGET_SHEET,
            "cell_ref": _cell_ref(row, TARGET_COLUMN),
            "value": validate_result(category),
        }
    ]
    if not link or not card_index:
        return updates

    for sheet_name, card_row, card_col in card_index.get(link, []):
        updates.append(
            {
                "sheet_name": sheet_name,
                "cell_ref": _cell_ref(card_row, card_col),
                "value": validate_result(category),
            }
        )
    return updates


def _apply_updates_to_workbook(wb, updates: list[CellUpdate]):
    for update in updates:
        ws = wb[update["sheet_name"]]
        ws[update["cell_ref"]] = update["value"]


def _sheet_xml_paths(zf: ZipFile) -> dict[str, str]:
    workbook = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

    rel_map: dict[str, str] = {}
    for rel in rels:
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rel_id and target:
            rel_map[rel_id] = target

    sheet_paths: dict[str, str] = {}
    for sheet in workbook.findall(f".//{{{XML_NS_MAIN}}}sheet"):
        sheet_name = sheet.attrib.get("name")
        rel_id = sheet.attrib.get(f"{{{XML_NS_REL}}}id")
        if not sheet_name or not rel_id or rel_id not in rel_map:
            continue
        target = rel_map[rel_id].lstrip("/")
        if not target.startswith("xl/"):
            target = f"xl/{target}"
        sheet_paths[sheet_name] = target
    return sheet_paths


def _set_inline_string(cell: ET.Element, value: str):
    for child in list(cell):
        if child.tag in {f"{{{XML_NS_MAIN}}}v", f"{{{XML_NS_MAIN}}}is"}:
            cell.remove(child)
    cell.set("t", "inlineStr")
    is_elem = ET.SubElement(cell, f"{{{XML_NS_MAIN}}}is")
    text_elem = ET.SubElement(is_elem, f"{{{XML_NS_MAIN}}}t")
    text_elem.text = value
    if value != value.strip() or "\n" in value:
        text_elem.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")


def _patch_sheet_xml(xml_bytes: bytes, updates: dict[str, str]) -> bytes:
    root = ET.fromstring(xml_bytes)
    sheet_data = root.find(f"{{{XML_NS_MAIN}}}sheetData")
    if sheet_data is None:
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    rows: dict[int, ET.Element] = {}
    cells: dict[str, ET.Element] = {}
    for row in sheet_data.findall(f"{{{XML_NS_MAIN}}}row"):
        row_ref = row.attrib.get("r")
        if row_ref and row_ref.isdigit():
            rows[int(row_ref)] = row
        for cell in row.findall(f"{{{XML_NS_MAIN}}}c"):
            cell_ref = cell.attrib.get("r")
            if cell_ref:
                cells[cell_ref] = cell

    for cell_ref, value in updates.items():
        cell = cells.get(cell_ref)
        if cell is None:
            row_number, _ = _split_cell_ref(cell_ref)
            row = rows.get(row_number)
            if row is None:
                row = ET.SubElement(
                    sheet_data, f"{{{XML_NS_MAIN}}}row", {"r": str(row_number)}
                )
                rows[row_number] = row
            cell = ET.SubElement(row, f"{{{XML_NS_MAIN}}}c", {"r": cell_ref})
            cells[cell_ref] = cell

        if cell.find(f"{{{XML_NS_MAIN}}}f") is not None:
            continue
        _set_inline_string(cell, value)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _layout_page_count(max_col: int) -> int:
    if max_col < PAGE_START_COLUMN:
        return 0
    return math.ceil((max_col - PAGE_START_COLUMN + 1) / PAGE_CONTENT_WIDTH)


def _shift_layout_column_index(col_idx: int) -> int:
    if col_idx < PAGE_START_COLUMN:
        return col_idx
    return col_idx + ((col_idx - PAGE_START_COLUMN) // PAGE_CONTENT_WIDTH)


def _layout_print_area_end_col(max_col: int) -> int:
    if max_col < PAGE_START_COLUMN:
        return PAGE_START_COLUMN
    return _shift_layout_column_index(max_col) + 1


def _shift_layout_marker_column(zero_based_col_idx: int) -> int:
    return _shift_layout_column_index(zero_based_col_idx + 1) - 1


def _shift_layout_cell_ref(cell_ref: str) -> str:
    row_idx, col_letters = _split_cell_ref(cell_ref)
    col_idx = column_index_from_string(col_letters)
    return f"{get_column_letter(_shift_layout_column_index(col_idx))}{row_idx}"


def _shift_layout_ref_token(token: str) -> str:
    if re.fullmatch(r"[A-Z]+\d+", token):
        return _shift_layout_cell_ref(token)
    if re.fullmatch(r"[A-Z]+\d+:[A-Z]+\d+", token):
        min_col, min_row, max_col, max_row = range_boundaries(token)
        if min_col is None or min_row is None or max_col is None or max_row is None:
            return token
        return (
            f"{get_column_letter(_shift_layout_column_index(min_col))}{min_row}:"
            f"{get_column_letter(_shift_layout_column_index(max_col))}{max_row}"
        )
    return token


def _shift_layout_sqref(value: str) -> str:
    return " ".join(_shift_layout_ref_token(token) for token in value.split())


def _shift_layout_spans(value: str) -> str:
    start, end = value.split(":", 1)
    return f"{_shift_layout_column_index(int(start))}:{_shift_layout_column_index(int(end))}"


def _coalesce_column_ranges(
    columns: list[tuple[int, dict[str, str]]],
) -> list[tuple[int, int, dict[str, str]]]:
    coalesced: list[tuple[int, int, dict[str, str]]] = []
    for col_idx, attrs in columns:
        if coalesced and coalesced[-1][1] + 1 == col_idx and coalesced[-1][2] == attrs:
            start, _, prev_attrs = coalesced[-1]
            coalesced[-1] = (start, col_idx, prev_attrs)
            continue
        coalesced.append((col_idx, col_idx, attrs))
    return coalesced


def _replace_cols_element(root: ET.Element, max_col: int):
    cols_tag = f"{{{XML_NS_MAIN}}}cols"
    col_tag = f"{{{XML_NS_MAIN}}}col"
    sheet_data_tag = f"{{{XML_NS_MAIN}}}sheetData"
    cols = root.find(cols_tag)
    expanded: list[tuple[int, dict[str, str]]] = []
    if cols is not None:
        for col in list(cols):
            min_col = int(col.attrib["min"])
            max_range_col = int(col.attrib["max"])
            attrs = {k: v for k, v in col.attrib.items() if k not in {"min", "max"}}
            for col_idx in range(min_col, max_range_col + 1):
                expanded.append((_shift_layout_column_index(col_idx), attrs))

    for page_idx in range(1, _layout_page_count(max_col) + 1):
        content_end = PAGE_START_COLUMN + (page_idx * PAGE_CONTENT_WIDTH) - 1
        pad_col = _shift_layout_column_index(content_end) + 1
        expanded.append(
            (
                pad_col,
                {"width": str(PAD_COLUMN_WIDTH), "customWidth": "1"},
            )
        )

    if not expanded:
        return

    expanded.sort(key=lambda item: item[0])
    new_cols = ET.Element(cols_tag)
    for start, end, attrs in _coalesce_column_ranges(expanded):
        col_attrs = {"min": str(start), "max": str(end), **attrs}
        ET.SubElement(new_cols, col_tag, col_attrs)

    if cols is not None:
        root.remove(cols)
    sheet_data = root.find(sheet_data_tag)
    if sheet_data is None:
        root.append(new_cols)
        return
    root.insert(list(root).index(sheet_data), new_cols)


def _ensure_worksheet_child(root: ET.Element, local_name: str, before_names: set[str]):
    tag = f"{{{XML_NS_MAIN}}}{local_name}"
    existing = root.find(tag)
    if existing is not None:
        return existing

    insert_at = len(root)
    for idx, child in enumerate(list(root)):
        child_name = child.tag.rsplit("}", 1)[-1]
        if child_name in before_names:
            insert_at = idx
            break
    new_child = ET.Element(tag)
    root.insert(insert_at, new_child)
    return new_child


def _max_sheet_column(root: ET.Element) -> int:
    dimension = root.find(f"{{{XML_NS_MAIN}}}dimension")
    if dimension is not None:
        ref = dimension.attrib.get("ref")
        if ref:
            _, _, max_col, _ = range_boundaries(ref)
            if max_col is not None:
                return max_col

    max_col = 1
    for cell in root.findall(f".//{{{XML_NS_MAIN}}}c"):
        cell_ref = cell.attrib.get("r")
        if not cell_ref:
            continue
        _, col_letters = _split_cell_ref(cell_ref)
        max_col = max(max_col, column_index_from_string(col_letters))
    return max_col


def _patch_layout_sheet_xml(xml_bytes: bytes) -> tuple[bytes, int]:
    root = ET.fromstring(xml_bytes)
    max_col = _max_sheet_column(root)

    _replace_cols_element(root, max_col)

    for elem in root.iter():
        if elem.tag == f"{{{XML_NS_MAIN}}}c":
            cell_ref = elem.attrib.get("r")
            if cell_ref:
                elem.set("r", _shift_layout_cell_ref(cell_ref))
        for attr_name in ("ref", "sqref", "activeCell", "topLeftCell"):
            value = elem.attrib.get(attr_name)
            if value:
                elem.set(attr_name, _shift_layout_sqref(value))
        spans = elem.attrib.get("spans")
        if spans:
            elem.set("spans", _shift_layout_spans(spans))

    print_options = _ensure_worksheet_child(
        root,
        "printOptions",
        {
            "pageMargins",
            "pageSetup",
            "headerFooter",
            "rowBreaks",
            "colBreaks",
            "drawing",
            "legacyDrawing",
        },
    )
    print_options.set("horizontalCentered", "1")
    print_options.set("verticalCentered", "1")

    page_margins = _ensure_worksheet_child(
        root,
        "pageMargins",
        {
            "pageSetup",
            "headerFooter",
            "rowBreaks",
            "colBreaks",
            "drawing",
            "legacyDrawing",
        },
    )
    left_margin = float(page_margins.attrib.get("left", "0.7"))
    page_margins.set("left", f"{left_margin + 0.2:g}")

    page_setup = _ensure_worksheet_child(
        root,
        "pageSetup",
        {"headerFooter", "rowBreaks", "colBreaks", "drawing", "legacyDrawing"},
    )
    page_setup.set("pageOrder", "overThenDown")

    return ET.tostring(root, encoding="utf-8", xml_declaration=True), max_col


def _normalize_zip_target(base_part: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return normpath(posix_join(dirname(base_part), target))


def _sheet_rels_path(sheet_path: str) -> str:
    file_name = sheet_path.rsplit("/", 1)[-1]
    return f"xl/worksheets/_rels/{file_name}.rels"


def _shift_drawing_xml(xml_bytes: bytes) -> bytes:
    root = ET.fromstring(xml_bytes)
    for marker_name in ("from", "to"):
        for marker in root.findall(f".//{{{XML_NS_XDR}}}{marker_name}"):
            col = marker.find(f"{{{XML_NS_XDR}}}col")
            if col is None or col.text is None:
                continue
            col.text = str(_shift_layout_marker_column(int(col.text)))
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _update_workbook_xml(
    xml_bytes: bytes,
    print_areas: dict[str, str],
    deleted_sheet_name: str | None,
) -> bytes:
    root = ET.fromstring(xml_bytes)
    sheets = root.find(f"{{{XML_NS_MAIN}}}sheets")
    if sheets is None:
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    removed_index: int | None = None
    sheet_elements = list(sheets)
    for idx, sheet in enumerate(sheet_elements):
        if sheet.attrib.get("name") != deleted_sheet_name:
            continue
        sheets.remove(sheet)
        removed_index = idx
        break

    remaining_sheets = list(sheets)
    new_index_by_name = {
        sheet.attrib["name"]: idx
        for idx, sheet in enumerate(remaining_sheets)
        if "name" in sheet.attrib
    }

    workbook_view = root.find(f".//{{{XML_NS_MAIN}}}workbookView")
    if workbook_view is not None and removed_index is not None:
        active_tab = workbook_view.attrib.get("activeTab")
        if active_tab and active_tab.isdigit():
            active_idx = int(active_tab)
            if active_idx == removed_index:
                workbook_view.set("activeTab", "0")
            elif active_idx > removed_index:
                workbook_view.set("activeTab", str(active_idx - 1))

    defined_names = root.find(f"{{{XML_NS_MAIN}}}definedNames")
    if defined_names is None and print_areas:
        defined_names = ET.Element(f"{{{XML_NS_MAIN}}}definedNames")
        calc_pr = root.find(f"{{{XML_NS_MAIN}}}calcPr")
        if calc_pr is None:
            root.append(defined_names)
        else:
            root.insert(list(root).index(calc_pr), defined_names)

    if defined_names is not None:
        for defined_name in list(defined_names):
            local_sheet_id = defined_name.attrib.get("localSheetId")
            if local_sheet_id is None or not local_sheet_id.isdigit():
                continue
            sheet_idx = int(local_sheet_id)
            if removed_index is not None:
                if sheet_idx == removed_index:
                    defined_names.remove(defined_name)
                    continue
                if sheet_idx > removed_index:
                    defined_name.set("localSheetId", str(sheet_idx - 1))
                    sheet_idx -= 1

            if (
                defined_name.attrib.get("name") == "_xlnm.Print_Area"
                and sheet_idx in new_index_by_name.values()
            ):
                defined_names.remove(defined_name)

        for sheet_name, print_area in print_areas.items():
            local_sheet_id = new_index_by_name.get(sheet_name)
            if local_sheet_id is None:
                continue
            defined_name = ET.SubElement(
                defined_names,
                f"{{{XML_NS_MAIN}}}definedName",
                {"name": "_xlnm.Print_Area", "localSheetId": str(local_sheet_id)},
            )
            defined_name.text = print_area

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _apply_layout_to_output_workbook(output_path: Path):
    temp_path = output_path.with_suffix(f"{output_path.suffix}.tmp")
    with ZipFile(output_path, "r") as source_zip:
        workbook_xml = source_zip.read("xl/workbook.xml")
        sheet_paths = _sheet_xml_paths(source_zip)
        patched_entries: dict[str, bytes] = {}
        print_areas: dict[str, str] = {}

        for sheet_name, sheet_path in sheet_paths.items():
            if sheet_name in {DELETE_SHEET_NAME, LAYOUT_EXCLUDED_SHEET}:
                continue
            patched_sheet, max_col = _patch_layout_sheet_xml(
                source_zip.read(sheet_path)
            )
            patched_entries[sheet_path] = patched_sheet
            print_areas[sheet_name] = (
                f"{quote_sheetname(sheet_name)}!$B$2:$"
                f"{get_column_letter(_layout_print_area_end_col(max_col))}$35"
            )

            rels_path = _sheet_rels_path(sheet_path)
            if rels_path not in source_zip.namelist():
                continue
            rels_root = ET.fromstring(source_zip.read(rels_path))
            for rel in rels_root:
                rel_type = rel.attrib.get("Type", "")
                target = rel.attrib.get("Target")
                if not target or not rel_type.endswith("/drawing"):
                    continue
                drawing_path = _normalize_zip_target(sheet_path, target)
                patched_entries[drawing_path] = _shift_drawing_xml(
                    source_zip.read(drawing_path)
                )

        patched_entries["xl/workbook.xml"] = _update_workbook_xml(
            workbook_xml,
            print_areas,
            None,
        )

        with ZipFile(temp_path, "w") as out_zip:
            out_zip.comment = source_zip.comment
            for item in source_zip.infolist():
                data = patched_entries.get(item.filename)
                if data is None:
                    data = source_zip.read(item.filename)
                out_zip.writestr(item, data)

    temp_path.replace(output_path)


def _apply_output_layout_to_workbook(wb):
    for ws in wb.worksheets:
        if ws.title in {DELETE_SHEET_NAME, LAYOUT_EXCLUDED_SHEET}:
            continue

        page_count = _layout_page_count(ws.max_column)
        for page_idx in range(page_count, 0, -1):
            insert_at = PAGE_START_COLUMN + (page_idx * PAGE_CONTENT_WIDTH)
            ws.insert_cols(insert_at)
            ws.column_dimensions[get_column_letter(insert_at)].width = PAD_COLUMN_WIDTH

        for page_idx in range(page_count):
            start_col = PAGE_START_COLUMN + (page_idx * PAGE_TOTAL_WIDTH)
            end_col = start_col + PAGE_CONTENT_WIDTH - 1
            for row_idx in range(PAGE_START_ROW, PAGE_END_ROW + 1):
                for col_idx in range(start_col, end_col + 1):
                    cell = ws.cell(row_idx, col_idx)
                    cell.border = Border(
                        left=THICK_BLACK_SIDE
                        if col_idx == start_col
                        else cell.border.left,
                        right=THICK_BLACK_SIDE
                        if col_idx == end_col
                        else cell.border.right,
                        top=THICK_BLACK_SIDE
                        if row_idx == PAGE_START_ROW
                        else cell.border.top,
                        bottom=THICK_BLACK_SIDE
                        if row_idx == PAGE_END_ROW
                        else cell.border.bottom,
                    )

        if page_count:
            ws.print_area = f"$B$2:${get_column_letter(PAGE_START_COLUMN + (page_count * PAGE_TOTAL_WIDTH) - 1)}$35"
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True
        ws.page_setup.pageOrder = "overThenDown"
        ws.page_margins.left = (ws.page_margins.left or 0.7) + 0.2


def save_workbook_with_preserved_media(
    source_path: Path, output_path: Path, updates: list[CellUpdate]
):
    selection = select_output_writer(source_path, updates)
    if not selection["supported"]:
        reason_code = selection["reason_code"] or "unsupported_unknown"
        detail = selection["detail"] or "workbook variant is outside supported contract"
        print(
            f"output_writer=unsupported policy={selection['policy']} reason_code={reason_code}",
            file=sys.stderr,
        )
        raise UnsupportedWorkbookVariantError(reason_code, detail)

    print(
        f"output_writer={selection['writer']} policy={selection['policy']}",
        file=sys.stderr,
    )

    _write_workbook_value_only_copy(source_path, output_path, updates)


def select_output_writer(
    source_path: Path, updates: Sequence[CellUpdate]
) -> OutputWriterSelection:
    contract_module = importlib.import_module("xlsx_template_contract")
    contract = contract_module.load_template_contract()
    contract_error_cls = contract_module.WorkbookContractError

    try:
        contract_module.validate_supported_workbook(source_path, contract)
        contract_module.enforce_writable_surface(
            cast(Sequence[Mapping[str, str]], updates), contract
        )
    except contract_error_cls as exc:
        return {
            "writer": "unsupported_variant",
            "policy": UNSUPPORTED_VARIANT_POLICY,
            "supported": False,
            "reason_code": str(exc.code),
            "detail": str(exc.detail),
        }

    return {
        "writer": PRIMARY_WRITER_ID,
        "policy": UNSUPPORTED_VARIANT_POLICY,
        "supported": True,
        "reason_code": None,
        "detail": None,
    }


def _write_workbook_value_only_copy(
    source_path: Path, output_path: Path, updates: Sequence[CellUpdate]
):

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_suffix(f"{output_path.suffix}.tmp")
    by_sheet: dict[str, dict[str, str]] = {}
    for update in updates:
        by_sheet.setdefault(update["sheet_name"], {})[update["cell_ref"]] = update[
            "value"
        ]

    with ZipFile(source_path, "r") as source_zip:
        sheet_paths = _sheet_xml_paths(source_zip)
        patched_entries: dict[str, bytes] = {}
        for sheet_name, sheet_updates in by_sheet.items():
            sheet_path = sheet_paths.get(sheet_name)
            if not sheet_path:
                continue
            patched_entries[sheet_path] = _patch_sheet_xml(
                source_zip.read(sheet_path), sheet_updates
            )

        with ZipFile(temp_path, "w") as out_zip:
            out_zip.comment = source_zip.comment
            for item in source_zip.infolist():
                data = patched_entries.get(item.filename)
                if data is None:
                    data = source_zip.read(item.filename)
                out_zip.writestr(item, data)

    temp_path.replace(output_path)


def save_workbook_with_preserved_media_legacy(
    source_path: Path, output_path: Path, updates: list[CellUpdate]
):
    save_workbook_with_preserved_media(source_path, output_path, updates)
    _apply_layout_to_output_workbook(output_path)


def collect_classification_updates(
    wb, api_key
) -> tuple[list[CellUpdate], dict[str, int]]:
    card_index = build_card_index(wb)
    targets = read_target_rows(wb)
    summary = {category: 0 for category in sorted(VALID_CATEGORIES)}
    updates: list[CellUpdate] = []

    total = len(targets)
    for idx, item in enumerate(targets, start=1):
        category, method = _classify_item_with_method(
            item["title"],
            item.get("link"),
            api_key,
        )
        updates.extend(
            _build_sheet_updates(item["row"], item.get("link"), category, card_index)
        )
        summary[category] += 1
        print(
            f"[{idx}/{total}] row={item['row']} ({method}) -> {category}",
            file=sys.stderr,
        )
    return updates, summary


def run_classification(wb, api_key, output_path, source_path: Path | None = None):
    updates, summary = collect_classification_updates(wb, api_key)
    _apply_updates_to_workbook(wb, updates)

    if output_path:
        if source_path is None:
            raise ValueError("source_path is required when output_path is provided")
        save_workbook_with_preserved_media(source_path, Path(output_path), updates)

    print("분류 완료 요약")
    for key in sorted(summary):
        print(f"{key}: {summary[key]}")
    return summary


def classify_workbook(wb, api_key, output_path, source_path: Path | None = None):
    return run_classification(
        wb, api_key=api_key, output_path=output_path, source_path=source_path
    )


INPUT_DIR = Path("input")


def _find_input_xlsx() -> Path | None:
    """Auto-detect a single xlsx file from the default input directory."""
    if not INPUT_DIR.is_dir():
        return None
    xlsx_files = sorted(INPUT_DIR.glob("*.xlsx"))
    if len(xlsx_files) == 1:
        return xlsx_files[0]
    return None


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="VOC workbook classifier")
    parser.add_argument(
        "input", nargs="?", help="Input xlsx path (default: auto-detect from input/)"
    )
    parser.add_argument("--output-dir", default="output/", help="Output directory")
    parser.add_argument("--dry-run", action="store_true", help="Skip API call")
    return parser


def main(argv=None) -> int:
    parser = _build_parser()
    args = parser.parse_args(argv)

    if not args.input:
        auto = _find_input_xlsx()
        if auto:
            args.input = str(auto)
            print(f"Auto-detected input: {auto}", file=sys.stderr)
        else:
            parser.print_help()
            return 0

    api_key = os.getenv("GITHUB_TOKEN")
    if not args.dry_run and not api_key:
        print("Error: GITHUB_TOKEN is required", file=sys.stderr)
        return 1

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: 파일을 찾을 수 없습니다: {input_path}", file=sys.stderr)
        return 1

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / input_path.name

    wb = openpyxl.load_workbook(input_path)
    if args.dry_run:
        targets = read_target_rows(wb)
        print(f"dry-run: {len(targets)} target rows")
        return 0

    run_classification(
        wb, api_key=api_key, output_path=output_path, source_path=input_path
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
